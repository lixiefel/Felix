"""
lite_excel.py — Produces a "lite" version of the audited Excel.

For paid clients who want to model what-ifs without seeing the model internals.

What it does:
- Opens the recalculated Excel
- Sets all sheets except OWNER_INPUTS and OWNER_RESULTS to "veryHidden"
  (these don't appear in Excel's unhide dropdown — power users need VBA to find them)
- Applies workbook structure protection with a password (read from secrets)
- Saves to a new file

What it does NOT do:
- Encrypt the workbook with strong AES (openpyxl can't do that). The Excel-level
  password just prevents casual snooping. A determined user can crack it.
  For higher security, use msoffcrypto-tool to wrap with AES post-export.
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook


# Sheets the client SHOULD see in the lite version.
VISIBLE_SHEETS = {"OWNER_INPUTS", "OWNER_RESULTS"}


def export_lite_excel(
    source_path: str,
    dest_path: str,
    protection_password: str | None = None,
) -> str:
    """
    Generates a lite Excel from source_path → dest_path.
    Returns dest_path on success. Raises on failure.

    `protection_password`: applied as workbook-structure protection.
    If None, no password is set (structure is still protected — sheets can't
    be unhidden or moved — but can be unprotected by anyone).
    """
    # Copy first so we never mutate the source
    shutil.copy(source_path, dest_path)

    wb = load_workbook(dest_path)

    # Hide internal sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name in VISIBLE_SHEETS:
            ws.sheet_state = "visible"
        else:
            # "veryHidden" means the sheet doesn't appear in Excel's Unhide dropdown
            ws.sheet_state = "veryHidden"

    # Workbook structure protection — prevents unhiding sheets and adding new ones
    if protection_password:
        wb.security.workbookPassword = protection_password
        wb.security.lockStructure = True
    else:
        wb.security.lockStructure = True

    # Reactivate a visible sheet so Excel opens to something the user can see
    for sheet_name in wb.sheetnames:
        if sheet_name in VISIBLE_SHEETS:
            wb.active = wb.sheetnames.index(sheet_name)
            break

    wb.save(dest_path)
    return dest_path
