"""
engine.py — MarginLab Excel pipeline
Writes inputs → recalculates via LibreOffice headless → reads results.
The Excel stays the canonical calculation engine. This file never touches formulas.
"""

import os
import shutil
import subprocess
import platform
import tempfile
import json
from dataclasses import dataclass, field
from typing import Optional
from pathlib import Path

from openpyxl import load_workbook

# ── paths ──────────────────────────────────────────────────────────────────────
TEMPLATE_PATH = Path(__file__).parent / "MarginLab_Pricing_Lab_v10.xlsx"

# ── LibreOffice recalc (mirrored from /mnt/skills/public/xlsx/scripts/recalc.py)
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def _get_soffice_env():
    """Return env dict for soffice subprocess (handles display/profile isolation)."""
    env = os.environ.copy()
    env.update({
        "HOME": str(Path.home()),
        "DISPLAY": ":99",
        "SAL_USE_VCLPLUGIN": "svp",
        "PYTHONPATH": "",
    })
    return env


def _setup_macro():
    macro_dir = os.path.expanduser(MACRO_DIR_LINUX)
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)
    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True
    # Init soffice profile if missing
    if not os.path.exists(macro_dir):
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True, timeout=15, env=_get_soffice_env()
        )
        os.makedirs(macro_dir, exist_ok=True)
    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalculate(filepath: str, timeout: int = 60) -> dict:
    """Run LibreOffice headless recalc on the file. Returns status dict."""
    if not _setup_macro():
        return {"ok": False, "error": "Failed to setup LibreOffice macro"}
    cmd = [
        "timeout", str(timeout),
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        str(Path(filepath).absolute()),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, env=_get_soffice_env())
    if result.returncode not in (0, 124):
        return {"ok": False, "error": result.stderr or "LibreOffice error"}
    # Count formula errors
    wb = load_workbook(filepath, data_only=True)
    errors = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#") and "!" in cell.value:
                    errors.append(f"{sname}!{cell.coordinate}={cell.value}")
    wb.close()
    return {"ok": len(errors) == 0, "errors": errors}


# ── data structures ────────────────────────────────────────────────────────────

@dataclass
class ItemInput:
    name: str
    category: str
    role: str
    cost: float
    price: float
    monthly_units: int
    comp1: Optional[float] = None
    comp2: Optional[float] = None
    comp3: Optional[float] = None


@dataclass
class SettingsInput:
    currency: str = "USD"
    round_to: float = 0.10
    ending: str = ".00"
    max_raise: float = 0.10
    max_cut: float = -0.05
    shr_high: float = 1.0
    shr_med: float = 0.5
    shr_low: float = 0.25
    demo_mode: str = "No"


@dataclass
class ItemResult:
    name: str
    action: str
    price_from: float
    price_to: float
    delta_pct: float
    delta_profit_mo: float
    confidence: str
    phase: str
    quadrant: str
    market: str
    narrative: str


@dataclass
class AuditResult:
    banner: Optional[str]          # None = clean, str = warning text
    monthly_lift: float
    lift_pct: float
    changes_count: str
    confidence: str
    best_item: str
    items: list[ItemResult]
    sens_conservative: float
    sens_baseline: float
    sens_optimistic: float
    sens_robust: str
    qa_hard_fails: int
    qa_soft_warns: int
    qa_info_obs: int
    qa_ready: str
    excel_path: str


# ── CATEGORIES and ROLES (must match CATEGORY_PRIORS and PSYCHOLOGY_RULES in Excel)
CATEGORIES = [
    "Coffee", "Specialty Drink", "Tea", "Pastry", "Dessert",
    "Sandwich/Food", "Breakfast", "Other",
]
ROLES = [
    "Traffic Driver", "Core", "Profit Driver",
    "Premium Anchor", "Signature", "Complement", "Other",
]
CURRENCIES = ["USD", "EUR", "GBP", "AUD", "CAD", "IDR", "JPY", "MYR", "PHP", "THB", "VND"]
ROUND_STEPS = [0.01, 0.05, 0.10, 0.50, 1.00, 500, 1000, 5000]
ENDINGS = [".00", ".50", ".90", ".95", ".99"]


# ── write inputs ───────────────────────────────────────────────────────────────

def write_inputs(wb, settings: SettingsInput, items: list[ItemInput]):
    """Populate SETTINGS, OWNER_INPUTS, COMPETITOR_BENCHMARK from form data."""
    # SETTINGS
    s = wb["SETTINGS"]
    s["B4"] = settings.currency
    s["B5"] = settings.round_to
    s["B6"] = settings.ending
    s["B9"] = settings.max_raise
    s["B10"] = settings.max_cut
    s["B11"] = settings.shr_high
    s["B12"] = settings.shr_med
    s["B13"] = settings.shr_low
    s["B16"] = settings.demo_mode

    # OWNER_INPUTS  — clear rows 6-35 first
    oi = wb["OWNER_INPUTS"]
    for r in range(6, 36):
        for c in range(2, 9):
            oi.cell(r, c).value = None

    for i, item in enumerate(items[:30]):
        r = 6 + i
        oi.cell(r, 2).value = item.name
        oi.cell(r, 3).value = item.category
        oi.cell(r, 4).value = item.role
        oi.cell(r, 5).value = item.cost
        oi.cell(r, 6).value = item.price
        oi.cell(r, 7).value = item.monthly_units

    # COMPETITOR_BENCHMARK — clear competitor price columns (D-F = cols 4-6, rows 9-38)
    cb = wb["COMPETITOR_BENCHMARK"]
    for r in range(9, 39):
        for c in [4, 5, 6]:
            cb.cell(r, c).value = None
    for i, item in enumerate(items[:30]):
        r = 9 + i
        if item.comp1 is not None: cb.cell(r, 4).value = item.comp1
        if item.comp2 is not None: cb.cell(r, 5).value = item.comp2
        if item.comp3 is not None: cb.cell(r, 6).value = item.comp3


# ── read results ───────────────────────────────────────────────────────────────

def _safe_float(v, default=0.0) -> float:
    try: return float(v)
    except (TypeError, ValueError): return default

def _safe_str(v, default="") -> str:
    return str(v) if v is not None else default

def _safe_int(v, default=0) -> int:
    try: return int(v)
    except (TypeError, ValueError): return default


def read_results(wb, excel_path: str) -> AuditResult:
    """Pull all display data from the recalculated workbook."""
    or_ = wb["OWNER_RESULTS"]
    qa = wb["QA_CHECKS"]
    sn = wb["SENSITIVITY"]

    # Headline
    banner_raw = or_["A3"].value
    banner = banner_raw if (isinstance(banner_raw, str) and banner_raw.strip()) else None
    monthly_lift = _safe_float(or_["A5"].value)
    lift_pct = _safe_float(or_["C5"].value)
    changes_count = _safe_str(or_["E5"].value)
    confidence = _safe_str(or_["G5"].value)
    best_item = _safe_str(or_["G7"].value)

    # Per-item table (rows 11-40, cols A-K)
    items = []
    for r in range(11, 41):
        name = or_.cell(r, 1).value
        if not name or str(name).strip() == "":
            continue
        items.append(ItemResult(
            name=_safe_str(name),
            action=_safe_str(or_.cell(r, 2).value),
            price_from=_safe_float(or_.cell(r, 3).value),
            price_to=_safe_float(or_.cell(r, 4).value),
            delta_pct=_safe_float(or_.cell(r, 5).value),
            delta_profit_mo=_safe_float(or_.cell(r, 6).value),
            confidence=_safe_str(or_.cell(r, 7).value),
            phase=_safe_str(or_.cell(r, 8).value),
            quadrant=_safe_str(or_.cell(r, 9).value),
            market=_safe_str(or_.cell(r, 10).value),
            narrative=_safe_str(or_.cell(r, 11).value),
        ))

    # Sensitivity — rows shifted based on when I built the sensitivity summary
    # Try rows 61-64 (sensitivity block may be at row 60-64)
    def try_cell(addr):
        try: return wb["OWNER_RESULTS"][addr].value
        except: return None

    # Pull directly from SENSITIVITY sheet totals for reliability
    sens_cons = _safe_float(sn["H40"].value)
    sens_base = _safe_float(sn["I40"].value)
    sens_opt  = _safe_float(sn["J40"].value)
    # Robust verdict
    sens_robust_raw = sn["H44"].value
    sens_robust = _safe_str(sens_robust_raw) if sens_robust_raw else (
        "YES — all positive" if (sens_cons >= 0 and sens_base >= 0 and sens_opt >= 0)
        else "NO — at least one scenario negative"
    )

    # QA summary
    hard_fails = _safe_int(qa["B37"].value)
    soft_warns = _safe_int(qa["B38"].value)
    info_obs   = _safe_int(qa["B39"].value)
    qa_ready   = _safe_str(qa["C37"].value)

    return AuditResult(
        banner=banner,
        monthly_lift=monthly_lift,
        lift_pct=lift_pct,
        changes_count=changes_count,
        confidence=confidence,
        best_item=best_item,
        items=items,
        sens_conservative=sens_cons,
        sens_baseline=sens_base,
        sens_optimistic=sens_opt,
        sens_robust=sens_robust,
        qa_hard_fails=hard_fails,
        qa_soft_warns=soft_warns,
        qa_info_obs=info_obs,
        qa_ready=qa_ready,
        excel_path=excel_path,
    )


# ── main pipeline ──────────────────────────────────────────────────────────────

def run_audit(settings: SettingsInput, items: list[ItemInput]) -> tuple[AuditResult | None, str]:
    """
    Full pipeline: write → save → recalc → read.
    Returns (AuditResult, error_message). error_message is "" on success.
    """
    if not TEMPLATE_PATH.exists():
        return None, f"Template not found at {TEMPLATE_PATH}"

    # Copy template to a unique tmp file
    tmp_dir = Path(tempfile.mkdtemp())
    tmp_path = tmp_dir / "audit_run.xlsx"
    shutil.copy(TEMPLATE_PATH, tmp_path)

    try:
        wb = load_workbook(str(tmp_path))
        write_inputs(wb, settings, items)
        wb.save(str(tmp_path))
        wb.close()

        result = recalculate(str(tmp_path))
        if not result["ok"]:
            errs = result.get("errors", [])
            return None, f"Recalculation errors: {'; '.join(errs[:5])}"

        wb2 = load_workbook(str(tmp_path), data_only=True)
        audit = read_results(wb2, str(tmp_path))
        wb2.close()
        return audit, ""

    except Exception as e:
        return None, str(e)
